#! python3
# sendDuesReminders.py - スプレッドシートの支払い状況に基づきメールを送信

from email.message import EmailMessage
from smtplib import SMTP
import sys

import openpyxl

EMAIL_ADDRESS = sys.argv[1]
APP_PASSWORD  = sys.argv[2]

# スプレッドシートを開き最近の支払い状況を取得
wb = openpyxl.load_workbook('duesRecords.xlsx', data_only=True)
ws = wb['Sheet1']

last_col = ws.max_column
latest_month = ws.cell(row=1, column=last_col).value

# 会員の支払い状況を調べる
unpaid_members = {}
for ri in range(2, ws.max_row + 1):
    payment = ws.cell(row=ri, column=last_col).value
    if payment != 'paid':
        name = ws.cell(row=ri, column=1).value
        email = ws.cell(row=ri, column=2).value
        unpaid_members[name] = email

# メールアカウントにログインする
with SMTP('smtp.gmail.com', 587) as server:
    server.starttls()
    server.login(EMAIL_ADDRESS, APP_PASSWORD)

    # リマインダーメールを送信する
    for name, email in unpaid_members.items():
        text = f'''Dear {name},
Records show that you have not paid dues for {latest_month}. Please make this payment as soon as possible. Thank you!'''
        
        msg = EmailMessage()
        msg['From']    = EMAIL_ADDRESS
        msg['To']      = email
        msg['Subject'] = f'{latest_month} dues unpaid.'
        msg.set_content(text)
        
        print(f'メール送信中 {email}...')
        sendmail_status = server.send_message(msg)

        if sendmail_status != {}:
            print(f'{email}へメール送信中に問題が起こりました: {sendmail_status}')
