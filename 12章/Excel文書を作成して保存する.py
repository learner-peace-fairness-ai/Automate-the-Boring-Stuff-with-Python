from openpyxl import Workbook

book = Workbook()
print(book.sheetnames)

sheet = book.active
print(sheet.title)

sheet.title = 'Spam Bacon Eggs Sheet'
print(book.sheetnames)
