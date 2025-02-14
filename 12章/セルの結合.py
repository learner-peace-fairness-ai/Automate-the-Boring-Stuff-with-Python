from openpyxl import Workbook

book = Workbook()
sheet = book.active

sheet.merge_cells('A1:D3')
sheet['A1'] = 'Twelve cells merged together'

sheet.merge_cells('C5:D5')
sheet['C5'] = 'Two merged cells'

book.save('merged.xlsx')
