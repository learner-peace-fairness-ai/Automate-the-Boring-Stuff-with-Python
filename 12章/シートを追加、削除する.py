from openpyxl import Workbook

book = Workbook()
print(book.sheetnames)

book.create_sheet()
print(book.sheetnames)

book.create_sheet(index=0, title='First Sheet')
print(book.sheetnames)

book.create_sheet(index=2, title='Middle Sheet')
print(book.sheetnames)

book.remove(book['Middle Sheet'])
book.remove(book['Sheet1'])
print(book.sheetnames)
