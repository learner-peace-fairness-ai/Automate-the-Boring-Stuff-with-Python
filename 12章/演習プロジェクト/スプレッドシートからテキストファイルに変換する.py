# エクセルファイルの列の値をテキストファイルに書き込む
# Usage: python スプレッドシートからテキストファイルに変換する.py excel_file

import sys

import openpyxl

if len(sys.argv) != 2:
    print('Usage: python スプレッドシートからテキストファイルに変換する.py excel_file')
    sys.exit()

excel_filename = sys.argv[1]

book = openpyxl.load_workbook(excel_filename)
sheet = book.active

for sheet_name in book.sheetnames:
    for col in sheet.iter_cols():
        ci = col[0].column
        
        with open(f'{sheet.title}_{ci}', 'w', encoding='utf-8') as fw:
            for cell in col:
                fw.write(f'{cell.value}\n')
