# エクセルファイルに空行を挿入する
# Usage: python blankRowInserter.py 空行の開始行 空行数 エクセルファイル

import sys

import openpyxl
from openpyxl import Workbook


def remove_all_sheets(book):
    for sheet_name in book.sheetnames:
        sheet = book[sheet_name]
        book.remove(sheet)


if len(sys.argv) != 4:
    print('Usage: python blankRowInserter.py 空行の開始行 空行数 エクセルファイル')
    sys.exit()

blank_line_start = int(sys.argv[1])
blank_line_count = int(sys.argv[2])
excel_filename = sys.argv[3]

src_book = openpyxl.load_workbook(excel_filename)
new_book = Workbook()

remove_all_sheets(new_book)

for sheet_name in src_book.sheetnames:
    src_sheet = src_book[sheet_name]
    new_sheet = new_book.create_sheet(title=sheet_name)

    ri = 1
    for row in src_sheet.iter_rows(min_row=1, max_row=blank_line_start - 1):
        for cell in row:
            ci = cell.column
            new_sheet.cell(row=ri, column=ci).value = cell.value
        ri += 1

    ri += blank_line_count
    
    for row in src_sheet.iter_rows(min_row=blank_line_start):
        for cell in row:
            ci = cell.column
            new_sheet.cell(row=ri, column=ci).value = cell.value
        ri += 1

src_book.close()
new_book.save(excel_filename)
