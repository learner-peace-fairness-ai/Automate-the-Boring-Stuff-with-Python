# N×N の掛け算の表をExcelファイルに保存する
# Usage: multiplicationTable.py N
# 1行とA列は太字

import sys

from openpyxl import Workbook
from openpyxl.styles import Font

FILE_NAME = 'multiplicationTable.xlsx'
ROW_LABEL_COLUMN_INDEX = 1
COLUMN_LABEL_ROW_INDEX = 1


def set_row_label(n, sheet, font):
    LABEL_ri = COLUMN_LABEL_ROW_INDEX
    LABEL_ci = ROW_LABEL_COLUMN_INDEX
    
    ri = 1
    for col in sheet.iter_cols(min_row=LABEL_ri + 1, max_row=LABEL_ri + n, min_col=LABEL_ci, max_col=LABEL_ci):
        for cell in col:
            cell.value = ri
            cell.font = font
            ri += 1


def set_column_label(n, sheet, font):
    LABEL_ri = COLUMN_LABEL_ROW_INDEX
    LABEL_ci = ROW_LABEL_COLUMN_INDEX

    ci = 1
    for row in sheet.iter_rows(min_col=LABEL_ci + 1, max_col=LABEL_ci + n, min_row=LABEL_ri, max_row=LABEL_ri):
        for cell in row:
            cell.value = ci
            cell.font = font
            ci += 1


if len(sys.argv) != 2:
    print('Usage: multiplicationTable.py N')
    sys.exit()
n = int(sys.argv[1])

book = Workbook()
sheet = book.active
bold_font = Font(bold=True)

set_row_label(n, sheet, bold_font)
set_column_label(n, sheet, bold_font)

LABEL_ri = COLUMN_LABEL_ROW_INDEX
LABEL_ci = ROW_LABEL_COLUMN_INDEX

ri = 1
for row in sheet.iter_rows(min_col=LABEL_ci + 1, min_row=LABEL_ri + 1, max_col=LABEL_ci + n, max_row=LABEL_ri + n):
    ci = 1
    for cell in row:
        cell.value = ri * ci
        ci += 1
    
    ri += 1

book.save(FILE_NAME)
