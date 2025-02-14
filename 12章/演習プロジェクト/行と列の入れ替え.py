# Usage: python 行と列の入れ替え.py エクセルファイル

import sys

import openpyxl
from openpyxl import Workbook


def remove_all_sheets(book):
    for sheet_name in book.sheetnames:
        sheet = book[sheet_name]
        book.remove(sheet)


if len(sys.argv) != 2:
    print('Usage: python 行と列の入れ替え.py エクセルファイル')
    sys.exit()

excel_filename = sys.argv[1]

src_book = openpyxl.load_workbook(excel_filename)
new_book = Workbook()

remove_all_sheets(new_book)

for sheet_name in src_book.sheetnames:
    src_sheet = src_book[sheet_name]
    new_sheet = new_book.create_sheet(title=sheet_name)

    for row in src_sheet.iter_rows():
        ri = row[0].row
        for cell in row:
            ci = cell.column
            new_sheet.cell(row=ci, column=ri).value = cell.value

src_book.close()
new_book.save(excel_filename)
