import openpyxl

book = openpyxl.load_workbook('example.xlsx')
sheet = book['Sheet1']

print(sheet['A1'])
print(sheet['A1'].value)

cell = sheet['B1']
print(cell.value)

print(f'行 {cell.row}, 列 {cell.column} は {cell.value}')
print(f'セル {cell.coordinate} は {cell.value}')
print(sheet['C1'].value)

print(f'{sheet.cell(row=1, column=2)}')
print(f'{sheet.cell(row=1, column=2).value}')
for i in range(1, 8, 2):
    print(i, sheet.cell(row=i, column=2).value)
