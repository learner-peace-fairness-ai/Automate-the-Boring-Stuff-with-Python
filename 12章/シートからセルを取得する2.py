import openpyxl

book = openpyxl.load_workbook('example.xlsx')
sheet = book['Sheet1']

print(sheet.max_row)
print(sheet.max_column)
