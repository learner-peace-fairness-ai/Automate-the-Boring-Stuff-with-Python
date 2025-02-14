import openpyxl

book = openpyxl.load_workbook('merged.xlsx')
sheet = book.active

sheet.unmerge_cells('A1:D3')
sheet.unmerge_cells('C5:D5')

book.save('unmerged.xlsx')
