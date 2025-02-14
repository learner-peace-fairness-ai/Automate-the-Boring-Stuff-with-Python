import openpyxl
from openpyxl.utils import column_index_from_string
from openpyxl.utils import get_column_letter

print(get_column_letter(1))
print(get_column_letter(2))
print(get_column_letter(27))
print(get_column_letter(900))

book = openpyxl.load_workbook('example.xlsx')
sheet = book['Sheet1']

print(get_column_letter(sheet.max_column))
print(column_index_from_string('A'))
print(column_index_from_string('AA'))
